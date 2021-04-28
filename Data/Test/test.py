from openpyxl import load_workbook
import csv

#wb_obj = load_workbook('10812_R3_With_time.xlsx')
wb_obj = load_workbook('10112_R2_With_time.xlsx.xlsx')
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column

#m_row = sheet_obj.max_row
m_row = 17
excel_data_arr = []

for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 13)
    # print(cell_obj.value)
    excel_data_arr.append(cell_obj.value)


actual_data_arr = []

with open('/Users/spoorthykanduri/Desktop/SE_project/TBOP/Data/Output/output.csv') as csv_file:
    csv_reader = csv.reader(csv_file)
    next(csv_reader)
    for row in csv_reader:
        value = row[2]
        value = int(value)
        actual_data_arr.append(value)
        

actual_data_arr = actual_data_arr[1:]
# print("actual", len(actual_data_arr))
# print("data", len(excel_data_arr))

print("actual_data_arr",actual_data_arr)
print("excel_data_arr",excel_data_arr)

arr_size = len(actual_data_arr)
count = 0

for i in range(1,arr_size):
    if(actual_data_arr[i] == excel_data_arr[i]):
        count+=1

percentage_accuracy = (count/arr_size)*100

print("accuracy of the code is :", percentage_accuracy)


from openpyxl import load_workbook
import csv

#wb_obj = load_workbook('10812_R3_With_time.xlsx')
wb_obj = load_workbook('Data/Evaluation/Eval_data/10112_R2_With_time.xlsx')
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column

m_row = sheet_obj.max_row
# m_row = 14
excel_data_arr = []

for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 13)
    # print(cell_obj.value)
    excel_data_arr.append(cell_obj.value)


actual_data_arr = []

with open('Data/Output/output.csv') as csv_file:
    csv_reader = csv.reader(csv_file)
    next(csv_reader)
    for row in csv_reader:
        key = row[1]
        value = row[2]
        value = int(value)
        actual_data_arr.append([key,value])
        

# actual_data_arr = actual_data_arr[1:]
print("actual", len(actual_data_arr))
print("data", len(excel_data_arr))

print("actual_data_arr",actual_data_arr)
print("excel_data_arr",excel_data_arr)

arr_size = len(actual_data_arr)
count = 0
correct = 0

for i in range(1,arr_size):
    if (int(int(actual_data_arr[i][0].split(':')[1])/5))%2 == 1:
            print(i,actual_data_arr[i][0])
            continue 
    for j in range(1,len(excel_data_arr)):

    # print(int(int(actual_data_arr[i][0].split(':')[1])/5)%2)
        
        if(actual_data_arr[i][1] == excel_data_arr[j]):
            count+=1
            correct += 1
        else:
            count +=1

percentage_accuracy = (correct/count)*100

print("accuracy of the code is :", percentage_accuracy)
