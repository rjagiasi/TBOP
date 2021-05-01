from openpyxl import load_workbook
import csv

#wb_obj = load_workbook('10812_R3_With_time.xlsx')
wb_obj = load_workbook('Data/Evaluation/91132_R2_With time.xlsx')
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column

m_row = sheet_obj.max_row
# m_row = 14
excel_data_arr = []
excel_time_arr = []


for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 13)
    time_obj = sheet_obj.cell(row = i, column = 1)
    # print(cell_obj.value)
    excel_data_arr.append(cell_obj.value)
    excel_time_arr.append(time_obj.value)


actual_data_arr = []

with open('Data/Output/output_91132_R3.csv') as csv_file:
    csv_reader = csv.reader(csv_file)
    next(csv_reader)
    for row in csv_reader:
        key = row[1]
        value = row[2]
        value = int(value)
        actual_data_arr.append([key,value])
        

# actual_data_arr = actual_data_arr[1:]
print("actual", len(actual_data_arr))
print("data", len(excel_data_arr))

print("actual_data_arr",actual_data_arr)
print("excel_data_arr",excel_data_arr)

arr_size = len(actual_data_arr)
total = 0
count = 0
correct = 0
spanish, english =0, 0
t_english = 0
ac_eng = 0
for i in range(0,arr_size):
    

    if (int(int(actual_data_arr[i][0].split(':')[1])/5))%2 == 1:
            # print(i,actual_data_arr[i][0])
            continue 
    # for j in range(0,len(excel_data_arr)):

    # print(int(int(actual_data_arr[i][0].split(':')[1])/5)%2)
    # print('Ours',i,actual_data_arr[i])
   
    #print("HH",actual_data_arr[i][1], excel_data_arr[count])
    if excel_data_arr[count] ==5:
        excel_data_arr[count] = 6

    if excel_data_arr[count] ==3:
        excel_data_arr[count] = 1
    if excel_data_arr[count] ==4:
        excel_data_arr[count] = 2
    if excel_data_arr[count] ==2:
        ac_eng+=1
    #print('Excel',excel_data_arr[count],actual_data_arr[i][1])
    print('Excel',excel_data_arr[count],actual_data_arr[i][1])
    if excel_data_arr[count]==2:
        t_english +=1
    if(actual_data_arr[i][1] == excel_data_arr[count]) :
        
        count +=1
        correct += 1
        total +=1
        if actual_data_arr[i][1] == 1:
            spanish += 1
        else:
            english += 1
    
    else:
        if excel_data_arr[count] not in [3,4,6]:
            count +=1
            total +=1
        elif excel_data_arr[count]==6:
            print("KK")
            count +=1
            total +=1
            correct +=1
        else:
            count +=1
            # total+=1
            #total += 1
    
    if count== len(excel_data_arr):
        break

    
print(correct)
print(count,total, i)
print(spanish, english)
print("english accuracy:", 100*english/t_english )
print("Percentage english predicted:", 100*english/total)
print("Actual english:",100*ac_eng/total )
percentage_accuracy = (correct/total)*100

print("accuracy of the code is :", percentage_accuracy)
